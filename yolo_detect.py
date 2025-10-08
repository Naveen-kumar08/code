import cv2
import os
from datetime import datetime
from ultralytics import YOLO
from openpyxl import Workbook, load_workbook

# ---------------------- YOLO Model ----------------------
model_path = r"C:\Users\Admin\Documents\YOLO\my_model\train\weights\best.pt"
model = YOLO(model_path)
labels = model.names

# ---------------------- Folder Setup ----------------------
image_folder = "Captured_Images"
os.makedirs(image_folder, exist_ok=True)

# ---------------------- Excel Setup ----------------------
excel_file = 'detected_levels.xlsx'
if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Detections"
    ws.append(["S.No", "Timestamp", "Captured Image", "Detected Level",
               "Confidence (%)", "X", "Y", "Width", "Height"])
    wb.save(excel_file)
else:
    wb = load_workbook(excel_file)
    ws = wb.active

# Find next serial number automatically
serial_no = ws.max_row  # header row counts as 1

# ---------------------- Camera Setup ----------------------
cap = cv2.VideoCapture(0)
print("Press 'C' to capture, 'Q' to quit")

while True:
    ret, frame = cap.read()
    if not ret:
        print("Camera not detected!")
        break

    # Run YOLO detection
    results = model(frame, verbose=False)
    detections = results[0].boxes

    # Draw bounding boxes on frame
    for box in detections:
        x1, y1, x2, y2 = map(int, box.xyxy[0])
        conf = float(box.conf[0])
        cls = int(box.cls[0])
        label = f"{labels[cls]} {conf:.2f}"
        cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 255, 0), 2)
        cv2.putText(frame, label, (x1, y1 - 10),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)

    # Display
    cv2.imshow("YOLO Bottle Level Detection", frame)

    key = cv2.waitKey(1) & 0xFF

    # Quit
    if key == ord('q') or key == ord('Q'):
        break

    # Capture
    elif key == ord('c') or key == ord('C'):
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        serial_no += 1
        img_name = f"{serial_no:03d}capture{timestamp}.png"
        img_path = os.path.join(image_folder, img_name)
        cv2.imwrite(img_path, frame)
        print(f"[{serial_no}] Image saved in {img_path}")

        detected_level = "Unknown"
        confidence_value = 0
        x = y = w = h = 0

        if len(detections) > 0:
            best_det = max(detections, key=lambda d: d.conf.item())
            classidx = int(best_det.cls.item())
            classname = labels[classidx].lower()
            confidence_value = round(best_det.conf.item() * 100, 2)

            # Get bounding box
            x1, y1, x2, y2 = map(int, best_det.xyxy[0])
            w = x2 - x1
            h = y2 - y1
            x = x1
            y = y1

            # Determine detected level
            if "full" in classname:
                detected_level = "Full level"
            elif "half" in classname:
                detected_level = "Half level"
            elif "empty" in classname:
                detected_level = "Empty level"
            else:
                detected_level = classname.capitalize()

        # Log to Excel (store only file name)
        ws.append([serial_no, timestamp, img_name, detected_level,
                   confidence_value, x, y, w, h])
        wb.save(excel_file)

        print(f"Logged: {detected_level} ({confidence_value}%) at [{x}, {y}, {w}, {h}]")

# ---------------------- Cleanup ----------------------
cap.release()
cv2.destroyAllWindows()
wb.save(excel_file)
print("Detection session ended. Excel saved successfully.")
